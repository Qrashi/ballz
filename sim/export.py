from datetime import datetime
from openpyxl import Workbook, worksheet, utils

import sim.data


# Gather data and export it into an excel file
def export_excel():
    """
    Export data into an excel file
    :return:
    """
    print("O Starting export", end="")
    workbook = Workbook()
    del workbook["Sheet"]
    workbook.create_sheet("overview")
    workbook.create_sheet("performance")
    workbook.create_sheet("all data")

    simdata = workbook["overview"]
    simdata.sheet_properties.tabColor = "c6b04d"

    simdata.cell(row=1, column=1).value = "Total iterations"
    simdata.column_dimensions[utils.get_column_letter(1)].width = len("Total iterations")
    simdata.cell(row=2, column=1).value = sim.iteration

    simdata.cell(row=1, column=2).value = "Total time [s]"
    simdata.column_dimensions[utils.get_column_letter(2)].width = len("Total time [s]")
    simdata.cell(row=2, column=2).value = sim.loop.realtime

    simdata.cell(row=1, column=3).value = "average CPU time per iteration [s]"
    print("\r\033[K\rO Calculating average cpu time", end="")
    simdata.column_dimensions[utils.get_column_letter(3)].width = len("average CPU time per iteration [ns]")
    simdata.cell(row=2, column=3).value = sum(sim.data.perf_time) / len(sim.data.perf_time)

    simdata.cell(row=1, column=4).value = "Export date"
    simdata.cell(row=2, column=4).value = datetime.now().strftime("%d.%m.%Y %H:%M:%S")
    simdata.column_dimensions[utils.get_column_letter(4)].width = len(datetime.now().strftime("%d.%m.%Y %H:%M:%S"))

    simdata.cell(row=1, column=5).value = "scenario"
    simdata.column_dimensions[utils.get_column_letter(1)].width = len("scenaro")
    simdata.cell(row=2, column=5).value = sim.scenarios.selected()

    simdata = workbook["performance"]
    simdata.sheet_properties.tabColor = "d6af15"
    print("\r\033[K\rO Saving performance data", end="")
    simdata.cell(row=1, column=1).value = "cpu time [ns]"
    row = 2
    for perf_data in sim.data.perf_time:
        simdata.cell(row=row, column=1).value = perf_data
        row += 1

    alldata = workbook["all data"]
    alldata.sheet_properties.tabColor = "15d6d6"
    alldatacolumn = 2
    sheet_setup(workbook["all data"])

    for sceneobj in sim.scene.data:
        workbook.create_sheet(sceneobj.name)
        sheet = workbook[sceneobj.name]
        sheet.sheet_properties.tabColor = "802f8e"
        # sheet_setup(sheet)
        cell = 1
        for plot, info in sceneobj.data.items():
            if not info["export"]:
                # Dont export
                continue
            data = info["data"]
            print("\r\033[K\rO Saving " + plot + " from " + sceneobj.name, end="")
            sheet.cell(row=1, column=cell).value = plot
            alldata.cell(row=1, column=alldatacolumn).value = sceneobj.name + " - " + plot
            sheet.column_dimensions[utils.get_column_letter(cell)].width = len(plot)
            alldata.column_dimensions[utils.get_column_letter(alldatacolumn)].width = len(plot)
            row = 2
            for point in data:
                sheet.cell(row=row, column=cell).value = point
                alldata.cell(row=row, column=alldatacolumn).value = point
                row += 1
            cell += 1
            alldatacolumn += 1

    print("\r\033[K\rO Saving workbook to file", end="")
    workbook.save(f"exports/ballz_data_{int(round(datetime.now().timestamp()))}.xlsx")
    print("\r\033[K\rOK Export complete!")

def sheet_setup(sheet: worksheet):
    """
    Setup a sheet with the delta_t column
    :param sheet: The sheet to set up
    :return:
    """
    sheet.cell(row=1, column=1).value = "real time [s]"
    sheet.column_dimensions[utils.get_column_letter(1)].width = len("real time [s]")

    iteration = 2
    print("\r\033[K\rO Saving real time data", end="")
    for time in sim.data.realtime:
        sheet.cell(row=iteration, column=1).value = time
        iteration += 1
